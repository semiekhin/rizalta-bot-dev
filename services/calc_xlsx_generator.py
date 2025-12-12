#!/usr/bin/env python3
"""
Генератор Excel-файла с расчётом прибыли от апартамента RIZALTA
Записывает вычисленные значения (не формулы) для совместимости с просмотрщиками
"""

import sqlite3
from pathlib import Path
from typing import Optional, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import tempfile

BASE_DIR = Path(__file__).parent.parent
DB_PATH = BASE_DIR / "properties.db"


def get_lot_from_db(code: str) -> Optional[Dict]:
    """Получить лот из БД по коду"""
    if not DB_PATH.exists():
        return None
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()
    code_upper = code.strip().upper()
    table = str.maketrans({"А": "A", "В": "B", "Е": "E", "К": "K", "М": "M", "Н": "H", "О": "O", "Р": "P", "С": "S", "Т": "T"})
    code_latin = code_upper.translate(table)
    cursor.execute("SELECT code, area_m2, price_rub FROM units WHERE code = ? OR code = ? LIMIT 1", (code_upper, code_latin))
    row = cursor.fetchone()
    conn.close()
    if row:
        return {"code": row[0], "area": row[1], "price": row[2], "price_m2": int(row[2] / row[1])}
    return None


def get_lot_by_area(area: float) -> Optional[Dict]:
    """Получить лот из БД по площади"""
    if not DB_PATH.exists():
        return None
    conn = sqlite3.connect(str(DB_PATH))
    cursor = conn.cursor()
    cursor.execute("SELECT code, area_m2, price_rub FROM units WHERE area_m2 = ? LIMIT 1", (area,))
    row = cursor.fetchone()
    conn.close()
    if row:
        return {"code": row[0], "area": row[1], "price": row[2], "price_m2": int(row[2] / row[1])}
    return None


class ProfitCalculatorGenerator:
    """Генератор Excel-файла с расчётом прибыли"""
    
    # Базовые ставки аренды / эталонная площадь 26.8 м²
    RENT_PRICES_PER_M2 = [664.18, 723.88, 787.31, 858.21, 932.84, 1014.93, 1104.48, 1201.49]
    
    # Загрузка отеля по годам, %
    OCCUPANCY = [40, 60, 70, 70, 70, 70, 70, 70]
    
    # Дней в году (2028-2035)
    DAYS_PER_YEAR = [366, 365, 365, 365, 366, 365, 365, 365]
    
    def __init__(self, area: float, price_m2: float, expense_rate: float = 0.5, growth_rate: float = 0.2):
        self.area = area
        self.price_m2 = price_m2
        self.expense_rate = expense_rate
        self.growth_rate = growth_rate
        self.total_cost = round(area * price_m2)
        self._setup_styles()
        self._precalculate()
    
    def _setup_styles(self):
        """Настройка стилей"""
        thin = Side(style='thin')
        self.border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)
        self.font_red_bold = Font(bold=True, color="FF0000")
        self.font_blue = Font(color="0070C0")
        self.font_bold = Font(bold=True)
        self.align_center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.align_center = Alignment(horizontal='center', vertical='center')
    
    def _precalculate(self):
        """Предварительный расчёт всех значений"""
        # Годы без аренды (2025-2027)
        self.pre_rent_data = []
        cumulative_growth = 0
        accumulated_value = self.total_cost
        
        # 2025: рост 18%
        growth_2025 = round(self.total_cost * 0.18)
        cumulative_growth += growth_2025
        self.pre_rent_data.append({
            'year': 2025,
            'growth_profit': growth_2025,
            'cumulative': cumulative_growth,
            'growth_pct': round(growth_2025 * 100 / self.total_cost, 2) if self.total_cost > 0 else 0,
            'total_pct': round(cumulative_growth * 100 / self.total_cost, 2) if self.total_cost > 0 else 0
        })
        accumulated_value += growth_2025
        
        # 2026: рост 20%
        growth_2026 = round(accumulated_value * self.growth_rate)
        cumulative_growth += growth_2026
        self.pre_rent_data.append({
            'year': 2026,
            'growth_profit': growth_2026,
            'cumulative': cumulative_growth,
            'growth_pct': round(growth_2026 * 100 / self.total_cost, 2) if self.total_cost > 0 else 0,
            'total_pct': round(cumulative_growth * 100 / self.total_cost, 2) if self.total_cost > 0 else 0
        })
        accumulated_value += growth_2026
        
        # 2027: рост 20%
        growth_2027 = round(accumulated_value * self.growth_rate)
        cumulative_growth += growth_2027
        self.pre_rent_data.append({
            'year': 2027,
            'growth_profit': growth_2027,
            'cumulative': cumulative_growth,
            'growth_pct': round(growth_2027 * 100 / self.total_cost, 2) if self.total_cost > 0 else 0,
            'total_pct': round(cumulative_growth * 100 / self.total_cost, 2) if self.total_cost > 0 else 0
        })
        accumulated_value += growth_2027
        
        # Годы с арендой (2028-2035)
        self.rent_data = []
        cumulative_rent = 0
        total_rent_profit_for_avg = 0
        
        for i, year in enumerate(range(2028, 2036)):
            daily_rate = round(self.RENT_PRICES_PER_M2[i] * self.area)
            occupancy = self.OCCUPANCY[i]
            days = self.DAYS_PER_YEAR[i]
            
            annual_rent = round(days * daily_rate * occupancy / 100)
            expenses = round(annual_rent * self.expense_rate)
            rent_profit = round(annual_rent - expenses)
            cumulative_rent += rent_profit
            total_rent_profit_for_avg += rent_profit
            
            # Рост стоимости
            if i == 0:  # 2028 - половина от 20%
                growth_profit = round(accumulated_value * self.growth_rate / 2)
            else:  # 2029-2035 - 8.8%
                growth_profit = round(accumulated_value * 0.088)
            
            accumulated_value += growth_profit
            cumulative_growth += growth_profit
            
            rent_pct = round(rent_profit * 100 / self.total_cost, 2) if self.total_cost > 0 else 0
            growth_pct = round(growth_profit * 100 / self.total_cost, 2) if self.total_cost > 0 else 0
            total_pct = round(rent_pct + growth_pct, 2)
            
            self.rent_data.append({
                'year': year,
                'daily_rate': daily_rate,
                'occupancy': occupancy,
                'annual_rent': annual_rent,
                'expenses': expenses,
                'rent_profit': rent_profit,
                'cumulative_rent': cumulative_rent,
                'growth_profit': growth_profit,
                'cumulative_total': cumulative_rent + cumulative_growth,
                'rent_pct': rent_pct,
                'growth_pct': growth_pct,
                'total_pct': total_pct
            })
        
        # Итоговые значения
        self.total_annual_rent = sum(d['annual_rent'] for d in self.rent_data)
        self.total_expenses = sum(d['expenses'] for d in self.rent_data)
        self.total_rent_profit = cumulative_rent
        self.total_growth_profit = cumulative_growth
        self.total_profit = self.rent_data[-1]['cumulative_total'] if self.rent_data else 0
        
        # Средняя прибыль за все 11 лет
        all_total_pcts = [d['total_pct'] for d in self.pre_rent_data] + [d['total_pct'] for d in self.rent_data]
        self.avg_total_pct = round(sum(all_total_pcts) / 11, 2) if all_total_pcts else 0
        
        # Сроки окупаемости
        self.payback_years = round(self.total_cost / (self.total_profit / 11), 2) if self.total_profit > 0 else 0
        avg_rent_profit = total_rent_profit_for_avg / 8 if total_rent_profit_for_avg > 0 else 1
        self.payback_rent_years = round(self.total_cost / avg_rent_profit, 2) if avg_rent_profit > 0 else 0
    
    def generate(self, output_path: Optional[str] = None) -> str:
        """Генерирует Excel-файл"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Расчет прибыли"
        
        self._set_dimensions(ws)
        self._create_header(ws)
        self._create_params_section(ws)
        self._create_table_header(ws)
        self._create_pre_rent_years(ws)
        self._create_rent_years(ws)
        self._create_totals(ws)
        
        if output_path is None:
            output_path = tempfile.mktemp(suffix='.xlsx', prefix='profit_calc_')
        
        wb.save(output_path)
        return output_path
    
    def _set_dimensions(self, ws):
        """Устанавливает размеры колонок и строк"""
        col_widths = {
            'A': 8.66, 'B': 16, 'C': 18, 'D': 13, 'E': 21, 
            'F': 18, 'G': 18, 'H': 22, 'I': 20, 'J': 20, 'K': 15.55, 
            'L': 16.5, 'M': 16.5, 'N': 17.5
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
        
        row_heights = {
            1: 21, 2: 21, 3: 25, 4: 21, 5: 80, 6: 21, 
            7: 21, 8: 21, 9: 17, 10: 80, 11: 29, 12: 29, 
            13: 29, 14: 29, 15: 29, 16: 29, 17: 29, 18: 29, 
            19: 29, 20: 29, 21: 29, 22: 24
        }
        for row, height in row_heights.items():
            ws.row_dimensions[row].height = height
    
    def _create_header(self, ws):
        """Создаёт заголовок"""
        ws['C3'] = 'Расчет прибыли'
        ws['C3'].font = Font(bold=True, size=14)
    
    def _create_params_section(self, ws):
        """Создаёт секцию с параметрами (строки 5-6)"""
        headers = {
            'B': 'Площадь, м2',
            'C': 'Цена за м2',
            'D': 'Стоимость, руб',
            'F': 'Расходы, %',
            'G': 'Прибыль от роста стоимости кв.м. апартамента в год, %',
            'H': 'Срок окупаемости апартамента, лет',
            'I': 'Срок окупаемости апартамента (от сдачи), лет'
        }
        
        for col, val in headers.items():
            cell = ws[f'{col}5']
            cell.value = val
            cell.font = self.font_bold
            cell.alignment = self.align_center_wrap
            cell.border = self.border_thin
        
        ws.merge_cells('D5:E5')
        ws.merge_cells('D6:E6')
        
        # Входные данные — ЗНАЧЕНИЯ вместо формул
        ws['B6'] = self.area
        ws['B6'].border = self.border_thin
        ws['B6'].alignment = self.align_center
        
        ws['C6'] = self.price_m2
        ws['C6'].border = self.border_thin
        ws['C6'].alignment = self.align_center
        ws['C6'].number_format = '#,##0'
        
        ws['D6'] = self.total_cost  # Было: '=B6*C6'
        ws['D6'].border = self.border_thin
        ws['D6'].alignment = self.align_center
        ws['D6'].number_format = '#,##0'
        
        ws['E6'].border = self.border_thin
        
        ws['F6'] = self.expense_rate
        ws['F6'].border = self.border_thin
        ws['F6'].alignment = self.align_center
        ws['F6'].number_format = '0%'
        
        ws['G6'] = self.growth_rate
        ws['G6'].font = self.font_red_bold
        ws['G6'].border = self.border_thin
        ws['G6'].alignment = self.align_center
        ws['G6'].number_format = '0%'
        
        ws['H6'] = self.payback_years  # Было: '=ROUND(D6/(I22/11),2)'
        ws['H6'].font = self.font_red_bold
        ws['H6'].border = self.border_thin
        ws['H6'].alignment = self.align_center
        ws['H6'].number_format = '0.00'
        
        ws['I6'] = self.payback_rent_years  # Было: '=ROUND(D6/(SUM(G11:G21)/9),2)'
        ws['I6'].font = self.font_red_bold
        ws['I6'].border = self.border_thin
        ws['I6'].alignment = self.align_center
        ws['I6'].number_format = '0.00'
    
    def _create_table_header(self, ws):
        """Создаёт заголовок основной таблицы (строка 10)"""
        headers = {
            'B': 'Год',
            'C': 'Стоимость сдачи номера в сутки, руб',
            'D': 'Загрузка отеля,%',
            'E': 'Стоимость сдачи номера в год, руб',
            'F': 'Расходы, руб',
            'G': 'Прибыль от сдачи апартамента, руб',
            'H': 'Накопительная прибыль от сдачи апартамента, руб',
            'I': 'Прибыль от роста стоимости кв.м. апартамента, руб',
            'J': 'Прибыль накопительно',
            'L': 'Прибыль от сдачи апартамента в год, %',
            'M': 'Прибыль роста цены от стоимости апартамента в год, %',
            'N': 'Общая прибыль от стоимости апартамента в год, %'
        }
        
        for col, val in headers.items():
            cell = ws[f'{col}10']
            cell.value = val
            cell.font = self.font_bold
            cell.alignment = self.align_center_wrap
            cell.border = self.border_thin
    
    def _create_pre_rent_years(self, ws):
        """Создаёт строки 11-13 (годы без аренды)"""
        for i, data in enumerate(self.pre_rent_data):
            row = 11 + i
            
            ws[f'B{row}'] = data['year']
            ws[f'B{row}'].border = self.border_thin
            ws[f'B{row}'].alignment = self.align_center
            
            # Пустые ячейки с рамкой
            for col in ['C', 'D', 'E', 'F', 'G', 'H', 'L']:
                ws[f'{col}{row}'].border = self.border_thin
            
            ws[f'I{row}'] = data['growth_profit']
            ws[f'I{row}'].border = self.border_thin
            ws[f'I{row}'].alignment = self.align_center
            ws[f'I{row}'].number_format = '#,##0'
            
            ws[f'J{row}'] = data['cumulative']
            ws[f'J{row}'].border = self.border_thin
            ws[f'J{row}'].alignment = self.align_center
            ws[f'J{row}'].number_format = '#,##0'
            
            ws[f'M{row}'] = data['growth_pct']
            ws[f'M{row}'].border = self.border_thin
            ws[f'M{row}'].alignment = self.align_center
            ws[f'M{row}'].number_format = '0.00'
            
            ws[f'N{row}'] = data['total_pct']
            ws[f'N{row}'].border = self.border_thin
            ws[f'N{row}'].alignment = self.align_center
            ws[f'N{row}'].number_format = '0.00'
    
    def _create_rent_years(self, ws):
        """Создаёт строки 14-21 (годы с арендой)"""
        for i, data in enumerate(self.rent_data):
            row = 14 + i
            
            ws[f'B{row}'] = data['year']
            ws[f'B{row}'].font = self.font_blue
            ws[f'B{row}'].border = self.border_thin
            ws[f'B{row}'].alignment = self.align_center
            
            ws[f'C{row}'] = data['daily_rate']
            ws[f'C{row}'].border = self.border_thin
            ws[f'C{row}'].alignment = self.align_center
            ws[f'C{row}'].number_format = '#,##0'
            
            ws[f'D{row}'] = data['occupancy']
            ws[f'D{row}'].border = self.border_thin
            ws[f'D{row}'].alignment = self.align_center
            
            ws[f'E{row}'] = data['annual_rent']
            ws[f'E{row}'].border = self.border_thin
            ws[f'E{row}'].alignment = self.align_center
            ws[f'E{row}'].number_format = '#,##0'
            
            ws[f'F{row}'] = data['expenses']
            ws[f'F{row}'].border = self.border_thin
            ws[f'F{row}'].alignment = self.align_center
            ws[f'F{row}'].number_format = '#,##0'
            
            ws[f'G{row}'] = data['rent_profit']
            ws[f'G{row}'].border = self.border_thin
            ws[f'G{row}'].alignment = self.align_center
            ws[f'G{row}'].number_format = '#,##0'
            
            ws[f'H{row}'] = data['cumulative_rent']
            ws[f'H{row}'].border = self.border_thin
            ws[f'H{row}'].alignment = self.align_center
            ws[f'H{row}'].number_format = '#,##0'
            
            ws[f'I{row}'] = data['growth_profit']
            ws[f'I{row}'].border = self.border_thin
            ws[f'I{row}'].alignment = self.align_center
            ws[f'I{row}'].number_format = '#,##0'
            
            ws[f'J{row}'] = data['cumulative_total']
            ws[f'J{row}'].border = self.border_thin
            ws[f'J{row}'].alignment = self.align_center
            ws[f'J{row}'].number_format = '#,##0'
            
            ws[f'L{row}'] = data['rent_pct']
            ws[f'L{row}'].border = self.border_thin
            ws[f'L{row}'].alignment = self.align_center
            ws[f'L{row}'].number_format = '0.00'
            
            ws[f'M{row}'] = data['growth_pct']
            ws[f'M{row}'].border = self.border_thin
            ws[f'M{row}'].alignment = self.align_center
            ws[f'M{row}'].number_format = '0.00'
            
            ws[f'N{row}'] = data['total_pct']
            ws[f'N{row}'].border = self.border_thin
            ws[f'N{row}'].alignment = self.align_center
            ws[f'N{row}'].number_format = '0.00'
    
    def _create_totals(self, ws):
        """Создаёт строку итогов (строка 22)"""
        ws['B22'] = 'Итого'
        ws['B22'].font = self.font_bold
        ws['B22'].border = self.border_thin
        ws['B22'].alignment = self.align_center
        
        # Пустые ячейки с границами
        for col in ['C', 'D', 'L', 'M']:
            ws[f'{col}22'].border = self.border_thin
        
        ws['E22'] = self.total_annual_rent
        ws['E22'].border = self.border_thin
        ws['E22'].alignment = self.align_center
        ws['E22'].number_format = '#,##0'
        
        ws['F22'] = self.total_expenses
        ws['F22'].border = self.border_thin
        ws['F22'].alignment = self.align_center
        ws['F22'].number_format = '#,##0'
        
        ws['G22'] = self.total_rent_profit
        ws['G22'].border = self.border_thin
        ws['G22'].alignment = self.align_center
        ws['G22'].number_format = '#,##0'
        
        ws['H22'] = self.total_rent_profit
        ws['H22'].border = self.border_thin
        ws['H22'].alignment = self.align_center
        ws['H22'].number_format = '#,##0'
        
        ws['I22'] = self.total_growth_profit
        ws['I22'].border = self.border_thin
        ws['I22'].alignment = self.align_center
        ws['I22'].number_format = '#,##0'
        
        ws['J22'] = self.total_profit
        ws['J22'].border = self.border_thin
        ws['J22'].alignment = self.align_center
        ws['J22'].number_format = '#,##0'
        
        ws['N22'] = self.avg_total_pct
        ws['N22'].border = self.border_thin
        ws['N22'].alignment = self.align_center
        ws['N22'].number_format = '0.00'


def generate_roi_xlsx(unit_code: str = None, area: float = None, output_dir: str = None) -> Optional[str]:
    """
    Генерирует Excel-файл с расчётом прибыли
    """
    lot = None
    if unit_code:
        lot = get_lot_from_db(unit_code)
    elif area:
        lot = get_lot_by_area(area)
    
    if not lot:
        print(f"[XLSX] Лот не найден: code={unit_code}, area={area}")
        return None
    
    generator = ProfitCalculatorGenerator(
        area=lot['area'],
        price_m2=lot['price_m2'],
        expense_rate=0.5,
        growth_rate=0.2
    )
    
    if output_dir is None:
        output_dir = tempfile.gettempdir()
    
    output_path = Path(output_dir) / f"ROI_{lot['code']}.xlsx"
    
    try:
        generator.generate(str(output_path))
        print(f"[XLSX] ✅ Создан: {output_path}")
        return str(output_path)
    except Exception as e:
        print(f"[XLSX] Ошибка: {e}")
        return None


if __name__ == "__main__":
    import sys
    code = sys.argv[1] if len(sys.argv) > 1 else "В200"
    generate_roi_xlsx(unit_code=code, output_dir="/tmp")
